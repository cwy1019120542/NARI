import poplib
import chardet
import email
import smtplib
import openpyxl
import datetime
import os
import uuid
import time
import shutil
import json
from traceback import print_exc
from ..models import MainConfig, ReceiveHistory
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
from ..extention import celery, Session, receive_logger, redis
from ..parameter_config import accept_file_type
from ..func_tools import get_match_dict, get_file_path, response, get_header_row, get_column_number, smtp_send_mail, send_multi_mail, to_xlsx, clean_file_name

def generate_log(main_config_id, level, message):
    getattr(receive_logger, level)(' '.join(["main_config_id", str(main_config_id), message]))
    redis.rpush(f"{main_config_id}_receive_log", time.strftime("%Y-%m-%d %H:%M:%S ") + message)

def decode_str(s):  # 字符编码转换
    value, charset = decode_header(s)[0]
    if charset:
        value = value.decode(charset)
    return value

def getTimeStamp(utcstr):
    format_str_list = [('%a, %d %b %Y %H:%M:%S +0800', 0), ('%a, %d %b %Y %H:%M:%S +0800 (CST)', 0), ('%a, %d %b %Y %H:%M:%S -0900', 17), ('%a, %d %b %Y %H:%M:%S -0900 (CST)', 17)]
    for format_str, add_hour in format_str_list:
        try:
            utcdatetime = datetime.datetime.strptime(utcstr, format_str)
        except Exception as error:
            continue
        utcdatetime += datetime.timedelta(hours=add_hour)
        localtimestamp = utcdatetime.timestamp()
        return localtimestamp

def get_column_data_list(sheet, field_data_list, header_row, insert_none=True):
    column_number_list = get_column_number()
    max_row = sheet.max_row - header_row
    field_data_list_test = [i.value for i in sheet[header_row]]
    column_data_list = []
    for field in field_data_list:
        if not field:
            column_data_list.append([None for i in range(max_row)])
        if field in field_data_list_test:
            field_index = field_data_list_test.index(field)
            column_number = column_number_list[field_index]
            column_data = [i.value for i in sheet[column_number][header_row:]]
            column_data_list.append(column_data)
        else:
            if insert_none:
                column_data_list.append([None for i in range(max_row)])
    return column_data_list

def decode_group(data):
    if data[-1][1]:
        decode_data = data[-1][0].decode(data[-1][1])
    else:
        decode_data = data[-1][0]
        if isinstance(decode_data, bytes):
            decode_data = decode_data.decode()
        if ' ' in decode_data:
            decode_data = decode_data.split(' ')[1][1:-1]
    decode_data = decode_data.strip().strip('<').strip('>')
    return decode_data

def get_target_group(target, match_dict):
    target_email = match_dict[target]
    target_group = [target, '|'.join(target_email)]
    return target_group

@celery.task(track_started=True)
def receive_mail(app_config_info, main_config_info):
    try:
        main_config_id = main_config_info['id']
        main_config_name = main_config_info["config_name"]
        result_excel_name = f"{main_config_name}.xlsx"
        redis.delete(f"{main_config_id}_receive_log")
        generate_log(main_config_id, "info", f"收邮件任务开始")
        config_files_dir = app_config_info['CONFIG_FILES_DIR']
        file_dir = os.path.join(config_files_dir, str(main_config_id), "receive_excel")
        if os.path.exists(file_dir):
            shutil.rmtree(file_dir)
        os.makedirs(file_dir)
        username = main_config_info['email']
        password = main_config_info['password']
        receive_config_info = main_config_info["receive_config_info"]
        host = receive_config_info['ip']
        port = receive_config_info['port']
        match_exists, match_dict = get_match_dict(config_files_dir, main_config_id)
        new_match_dict = {}
        target_list = []
        if not match_exists:
            generate_log(main_config_id, "info", f"未上传邮箱对应表")
        else:
            generate_log(main_config_id, "info", f"邮箱对应表存在")
            for key, value in match_dict.items():
                for from_email in value:
                    if from_email in new_match_dict:
                        new_match_dict[from_email].append(key)
                    else:
                        new_match_dict[from_email] = [key]
            target_list = list(match_dict.keys())
        target_subject = receive_config_info['subject']
        sheet_info = receive_config_info['sheet_info']
        read_start_timestamp = receive_config_info['read_start_timestamp']
        read_end_timestamp = receive_config_info['read_end_timestamp']
        is_remind = receive_config_info['is_remind']
        remind_subject = receive_config_info['remind_subject']
        remind_content = receive_config_info['remind_content']
        remind_ip = receive_config_info['remind_ip']
        remind_port = receive_config_info['remind_port']
        run_timestamp = time.time()
        generate_log(main_config_id, "info", f"开始收邮件")
        try:
            pop_obj = poplib.POP3(host)
            pop_obj.user(username)
            pop_obj.pass_(password)
        except Exception as error:
            generate_log(main_config_id, "critical", f"无法登陆收件邮箱 {error}")
            return "运行失败,无法登陆收件邮箱"
        resp, mail_list, octets = pop_obj.list()
        mail_count = len(mail_list)
        error_format_target_list = set()
        no_attachment_target_list = set()
        generate_log(main_config_id, "info", f"mail_count {mail_count}")
        history_list = []
        excel_data_list = []
        for i in range(mail_count, 0, -1):
            try:
                resp, lines, octets = pop_obj.retr(i)
            except Exception as error:
                generate_log(main_config_id, "error", f"邮件解析失败 {error}")
                continue
            msg_content = b'\r\n'.join(lines).decode("utf8", "ignore")
            msg = Parser().parsestr(msg_content)
            if not msg['from'] or not msg['subject'] or not msg['date']:
                generate_log(main_config_id, "info", f"信息缺失")
                continue
            try:
                from_email = decode_header(msg['from'])
                from_email = decode_group(from_email)
                subject_group = decode_header(msg['subject'])
                subject = decode_group(subject_group)
                msg_timestamp = getTimeStamp(msg['date'])
            except:
                generate_log(main_config_id, "error", f"subject parse fail")
                continue
            if not msg_timestamp:
                generate_log(main_config_id, "error", f"{from_email} {subject} {msg['date']} 时间无法解析")
                continue
            if msg_timestamp < read_start_timestamp:
                generate_log(main_config_id, "info", f"{from_email} {subject} {msg_timestamp} 小于开始时间")
                break
            if msg_timestamp > read_end_timestamp:
                generate_log(main_config_id, "info", f"{from_email} {subject} {msg_timestamp} 大于结束时间")
                continue
            target_group_str = from_email
            if match_exists:
                if from_email not in new_match_dict:
                    generate_log(main_config_id, "info", f"{from_email} 不在配置内的邮箱")
                    continue
                target_group = new_match_dict[from_email]
                target_group_str = "|".join(target_group)
                check_list = [i for i in target_group if i not in target_list]
                if check_list:
                    generate_log(main_config_id, "info", f"{from_email} {new_match_dict.get(from_email)} 重复处理")
                    continue
            for part in msg.walk():
                file_name = part.get_filename()
                if file_name:
                    h = email.header.Header(file_name)
                    dh = email.header.decode_header(h)
                    filename = dh[0][0]
                    if dh[0][1]:
                        filename = decode_str(str(filename, dh[0][1]))
                        if target_subject not in filename:
                            continue
                        if not filename.endswith(accept_file_type):
                            if match_exists:
                                for target in target_group:
                                    if target in target_list:
                                        error_format_target_list.add(target)
                            generate_log(main_config_id, "error", f"filename {filename} 格式不合法")
                            continue
                        generate_log(main_config_id, "info", f"filename {filename} 符合条件")
                        data = part.get_payload(decode=True)
                        new_file_name = clean_file_name(f'{target_group_str}_{filename}')
                        file_path = os.path.join(file_dir, new_file_name)
                        with open(file_path, 'wb') as f:
                            f.write(data)
                        if not filename.endswith('.xlsx'):
                            to_xlsx(file_path)
                        if match_exists:
                            for target in target_group:
                                if target in target_list:
                                    target_list.remove(target)
                                if target in no_attachment_target_list:
                                    no_attachment_target_list.remove(target)
                                if target in error_format_target_list:
                                    error_format_target_list.remove(target)
                            history_list.append(ReceiveHistory(email=from_email, target=target_group_str,
                                                               create_timestamp=run_timestamp,
                                                               main_config_id=main_config_id, is_success=True,
                                                               message="success"))
                            excel_data_list.append([target_group_str, from_email, "成功"])
                else:
                    if match_exists:
                        for target in target_group:
                            if target in target_list:
                                no_attachment_target_list.add(target)
        # pop_obj.quit()
        if match_exists:
            remind_email_list = []
            for no_attachment_target in no_attachment_target_list:
                email_group = '|'.join(match_dict[no_attachment_target])
                remind_email_list.extend(match_dict[no_attachment_target])
                history_list.append(ReceiveHistory(email=email_group, target=no_attachment_target, create_timestamp=run_timestamp, main_config_id=main_config_id, is_success=False, message="缺失附件"))
                excel_data_list.append([no_attachment_target, email_group, "失败", "缺失附件"])
            for error_format_target in error_format_target_list:
                email_group = '|'.join(match_dict[error_format_target])
                remind_email_list.extend(match_dict[error_format_target])
                history_list.append(ReceiveHistory(email=email_group, target=error_format_target, create_timestamp=run_timestamp, main_config_id=main_config_id, is_success=False, message="附件格式不合法"))
                excel_data_list.append([error_format_target, email_group, "失败", "附件格式不合法"])
            for target in target_list:
                if target not in no_attachment_target_list:
                    email_group = '|'.join(match_dict[target])
                    remind_email_list.extend(match_dict[target])
                    history_list.append(ReceiveHistory(email=email_group, target=target, create_timestamp=run_timestamp, main_config_id=main_config_id, is_success=False, message="未回复"))
                    excel_data_list.append([target, email_group, "失败", "未回复"])
            if is_remind:
                send_multi_mail(remind_ip, remind_port, username, password, remind_email_list,
                                remind_subject, remind_content)
        generate_log(main_config_id, "info", f"邮件收取完毕")
        sheet_info = json.loads(sheet_info)
        is_success, return_data = get_file_path(config_files_dir, main_config_id, "template_excel")
        template_path = False
        if is_success:
            template_path = return_data
        is_success, receive_excel_path_list = get_file_path(config_files_dir, main_config_id, "receive_excel", True)
        if not is_success or not receive_excel_path_list:
            generate_log(main_config_id, "error", f"未收取到指定邮件")
        else:
            result_dir = os.path.join(config_files_dir, str(main_config_id), "result_excel")
            if os.path.exists(result_dir):
                shutil.rmtree(result_dir)
            os.makedirs(result_dir)
            result_path = os.path.join(result_dir, result_excel_name)
            if not template_path:
                generate_log(main_config_id, "info", f"没有提供模板")
                result_list = []
                receive_excel_check = openpyxl.load_workbook(receive_excel_path_list[0], data_only=True)
                sheet_name_list = receive_excel_check.sheetnames
                receive_excel_check.close()
                for sheet_name in sheet_name_list:
                    field_data_list = []
                    for receive_excel_path in receive_excel_path_list[:]:
                        receive_excel = openpyxl.load_workbook(receive_excel_path, data_only=True)
                        receive_excel_sheet_name_list = receive_excel.sheetnames
                        if sheet_name not in receive_excel_sheet_name_list:
                            generate_log(main_config_id, "error", f"{receive_excel_path} 没有sheet页 {sheet_name}")
                            continue
                        sheet = receive_excel[sheet_name]
                        header_row = get_header_row(sheet, None)
                        field_data_list_test = [i.value for i in sheet[header_row] if i.value]
                        for field in field_data_list_test:
                            if field not in field_data_list:
                                field_data_list.append(field)
                        receive_excel.close()
                    all_sheet_data = []
                    for path_index, receive_excel_path in enumerate(receive_excel_path_list):
                        receive_excel = openpyxl.load_workbook(receive_excel_path, data_only=True)
                        receive_excel_sheet_name_list = receive_excel.sheetnames
                        if sheet_name not in receive_excel_sheet_name_list:
                            generate_log(main_config_id, "error", f"{receive_excel_path} 没有sheet页 {sheet_name}")
                            continue
                        sheet = receive_excel[sheet_name]
                        header_row = get_header_row(sheet, None)
                        if path_index == 0:
                            header_data = [[i.value for i in sheet[i]] for i in range(1, header_row)]
                            all_sheet_data.extend(header_data)
                            all_sheet_data.append(field_data_list)
                        column_data_list = get_column_data_list(sheet, field_data_list, header_row)
                        all_sheet_data.extend(list(zip(*column_data_list)))
                        receive_excel.close()
                    result_list.append(all_sheet_data)
                    generate_log(main_config_id, "info", f"表格数据生成完毕")
                work_book = openpyxl.Workbook(result_excel_name)
                for sheet_name, result in reversed(list(zip(sheet_name_list, result_list))):
                    sheet = work_book.create_sheet(sheet_name, 0)
                    for data in result:
                        sheet.append(data)
                work_book.save(result_path)
                work_book.close()
            else:
                generate_log(main_config_id, "info", f"提供模板表")
                shutil.copy(template_path, result_path)
                result_list = []
                result_excel = openpyxl.load_workbook(result_path, data_only=True)
                result_excel_sheet_name_list = result_excel.sheetnames
                for sheet_name, header_row, merge_field, fill_field in sheet_info:
                    if len(sheet_info) == 1:
                        sheet = result_excel[result_excel_sheet_name_list[0]]
                    else:
                        if sheet_name not in result_excel_sheet_name_list:
                            generate_log(main_config_id, "error", f"result_excel 没有sheet页 {sheet_name}")
                            continue
                        else:
                            sheet = result_excel[sheet_name]
                    header_row = get_header_row(sheet, header_row)
                    field_data_list = [i.value for i in sheet[header_row]]
                    check_row_list = [i.value for i in sheet[header_row+1] if i.value]
                    if len(check_row_list) > 2:
                        generate_log(main_config_id, "info", f"模板表有数据")
                        merge_field_list = [i for i in merge_field.split('|')]
                        fill_field_list = [i for i in fill_field.split('|')]
                        merge_fill_dict = {}
                        for receive_excel_path in receive_excel_path_list:
                            generate_log(main_config_id, "info", f"聚合 {os.path.split(receive_excel_path)[1]}")
                            receive_excel = openpyxl.load_workbook(receive_excel_path, data_only=True)
                            receive_excel_sheet_name_list = receive_excel.sheetnames
                            if len(sheet_info) == 1:
                                receive_sheet = receive_excel[receive_excel_sheet_name_list[0]]
                            else:
                                if sheet_name not in receive_excel_sheet_name_list:
                                    generate_log(main_config_id, "error", f"receive_excel {receive_excel_path} 没有sheet页 {sheet_name}")
                                    continue
                                else:
                                    receive_sheet = receive_excel[sheet_name]
                            merge_data_list = get_column_data_list(receive_sheet, merge_field_list, header_row, False)
                            merge_data_group = list(zip(*merge_data_list))
                            fill_data_list = get_column_data_list(receive_sheet, fill_field_list, header_row, False)
                            fill_data_group = list(zip(*fill_data_list))
                            for merge_group, fill_group in zip(merge_data_group, fill_data_group):
                                fill_data = merge_fill_dict.get(tuple(merge_group))
                                if not fill_data:
                                    merge_fill_dict[tuple(merge_group)] = fill_group
                                else:
                                    if fill_data.count(None) > fill_group.count(None):
                                        merge_fill_dict[tuple(merge_group)] = fill_group
                            receive_excel.close()
                        result_list.append(merge_fill_dict)
                    else:
                        generate_log(main_config_id, "info", f"模板表无数据")
                        all_sheet_data = []
                        for receive_excel_path in receive_excel_path_list:
                            generate_log(main_config_id, "info", f"聚合 {os.path.split(receive_excel_path)[1]}")
                            receive_excel = openpyxl.load_workbook(receive_excel_path, data_only=True)
                            receive_excel_sheet_name_list = receive_excel.sheetnames
                            if len(sheet_info) == 1:
                                receive_sheet = receive_excel[receive_excel_sheet_name_list[0]]
                            else:
                                if sheet_name not in receive_excel_sheet_name_list:
                                    generate_log(main_config_id, "error", f"receive_excel {receive_excel_path} 没有sheet页 {sheet_name}")
                                    continue
                                else:
                                    receive_sheet = receive_excel[sheet_name]
                            column_data_list = get_column_data_list(receive_sheet, field_data_list, header_row)
                            all_row_data = []
                            for row_data in zip(*column_data_list):
                                if len(row_data) - row_data.count(None) < 3:
                                    filter_row_data = [i for i in row_data if i and len(str(i)) > 6]
                                    if not filter_row_data:
                                        continue
                                all_row_data.append(row_data)
                            all_sheet_data.extend(all_row_data)
                            receive_excel.close()
                        result_list.append(all_sheet_data)
                generate_log(main_config_id, "info", f"数据聚合完毕")
                result_excel.close()
                for result_index, result in enumerate(result_list):
                    sheet_info[result_index].append(result)
                result_excel = openpyxl.load_workbook(result_path, data_only=True)
                for sheet_name, header_row, merge_field, fill_field, result in sheet_info:
                    if len(sheet_info) == 1:
                        sheet = result_excel[result_excel_sheet_name_list[0]]
                    else:
                        if sheet_name not in result_excel_sheet_name_list:
                            generate_log(main_config_id, "error", f"result_excel 没有sheet页 {sheet_name}")
                            continue
                        else:
                            sheet = result_excel[sheet_name]
                    header_row = get_header_row(sheet, header_row)
                    field_data_list = [i.value for i in sheet[header_row]]
                    column_number_list = get_column_number()
                    if isinstance(result, list):
                        for data_index, data in enumerate(result, start=1):
                            for column_index, single_data in enumerate(data):
                                column_number = column_number_list[column_index]
                                sheet[f"{column_number}{header_row+data_index}"].value = single_data
                    else:
                        merge_field_dict = {}
                        merge_field_list = [i for i in merge_field.split('|')]
                        fill_field_list = [i for i in fill_field.split('|')]
                        for row_index, row_data in enumerate(list(sheet.values)[header_row:], start=header_row+1):
                            key_list = []
                            for merge_field in merge_field_list:
                                merge_field_index = field_data_list.index(merge_field)
                                key_list.append(row_data[merge_field_index])
                            merge_field_dict[tuple(key_list)] = row_index
                        for fill_field_in_index, fill_field in enumerate(fill_field_list):
                            fill_field_index = field_data_list.index(fill_field)
                            column_number = column_number_list[fill_field_index]
                            for merge_field_group, row_index in merge_field_dict.items():
                                fill_value_list = result.get(merge_field_group)
                                if fill_value_list:
                                    fill_value = fill_value_list[fill_field_in_index]
                                else:
                                    fill_value = None
                                sheet[f"{column_number}{row_index}"].value = fill_value
                result_excel.save(result_path)
                result_excel.close()
        session = Session()
        session.query(ReceiveHistory).filter_by(main_config_id=main_config_id).delete()
        if match_exists:
            session.add_all(history_list)
        session.query(MainConfig).filter_by(id=main_config_id).update({"run_timestamp": run_timestamp})
        session.commit()
        session.close()
        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet.append(["序号", "单位", "邮箱", "状态", "原因"])
        for data_index, excel_data in enumerate(excel_data_list, start=1):
            sheet.append([data_index] + excel_data)
        status_excel_path = os.path.join(config_files_dir, str(main_config_id), "收件状态表.xlsx")
        workbook.save(status_excel_path)
        workbook.close()
        generate_log(main_config_id, "info", f"运行成功")
        return "运行成功"
    except Exception as error:
        generate_log(main_config_id, "critical", f"运行失败 {print_exc()}")
        return "运行失败,未知错误"
