import os
import shutil
import re
import smtplib
import openpyxl
import time
import openpyxl
from traceback import print_exc
from ..models import MainConfig, SendHistory
from ..extention import celery, Session, send_logger, redis
from ..func_tools import response, get_file_path, get_match_dict, check_dir, smtp_send_mail, get_header_row, clean_file_name

def generate_log(main_config_id, level, message):
    getattr(send_logger, level)(' '.join(["main_config_id", str(main_config_id), message]))
    redis.rpush(f"{main_config_id}_send_log", time.strftime("%Y-%m-%d %H:%M:%S ") + message)

def safe_send(smtp_obj, email_list, sender, subject, content, ip, port, password, key, main_config_id, run_timestamp, attachment_list=[]):
    history_list = []
    excel_data_list = []
    for email in email_list:
        send_success = False
        for i in range(5):
            try:
                smtp_send_mail(smtp_obj, sender, email, subject, content, attachment_list)
            except Exception as error:
                if i == 0:
                    # smtp_obj.quit()
                    smtp_obj = smtplib.SMTP(ip, port)
                    smtp_obj.login(sender, password)
            else:
                send_success = True
                break
        if not send_success:
            history_list.append(SendHistory(target=key, main_config_id=main_config_id, email=email,
                                            create_timestamp=run_timestamp, is_success=False,
                                            message="发送失败"))
            excel_data_list.append([key, email, "失败"])
            generate_log(main_config_id, "error",
                         f"main_config_id {main_config_id} {email} 发送失败")
        else:
            history_list.append(SendHistory(target=key, main_config_id=main_config_id, email=email,
                                            create_timestamp=run_timestamp, is_success=True,
                                            message="success"))
            excel_data_list.append([key, email, "成功"])
            generate_log(main_config_id, "info", f"main_config_id {main_config_id} {email} 发送成功")
    return history_list, excel_data_list

@celery.task(track_started=True)
def send_mail(app_config_info, main_config_info):
    try:
        main_config_id = main_config_info['id']
        redis.delete(f"{main_config_id}_send_log")
        generate_log(main_config_id, "info", f"main_config_id {main_config_id} 发邮件任务开始")
        config_files_dir = app_config_info['CONFIG_FILES_DIR']
        is_success, return_data = get_match_dict(config_files_dir, main_config_id)
        if not is_success:
            generate_log(main_config_id, "critical", f"main_config_id {main_config_id} 缺失邮箱对应表")
            return "运行失败,缺失邮箱对应表"
        match_dict = return_data
        total_target_list = list(match_dict.keys())
        send_config_info = main_config_info["send_config_info"]
        subject = send_config_info['subject']
        content = send_config_info['content']
        split_type = send_config_info["split_type"]
        sheet_list = send_config_info['sheet'].split('|')
        field_row_list = [int(i) if i else None for i in send_config_info['field_row'].split('|')]
        is_split_list = [int(i) if i else 0 for i in send_config_info['is_split'].split('|')]
        split_field = send_config_info['split_field']
        ip = send_config_info['ip']
        port = send_config_info['port']
        sender = main_config_info['email']
        password = main_config_info['password']
        try:
            smtp_obj = smtplib.SMTP(ip, port)
            smtp_obj.login(sender, password)
        except Exception as error:
            generate_log(main_config_id, "critical", f"main_config_id {main_config_id} 无法登陆发件邮箱 {error}")
            return "运行失败,无法登陆发件邮箱"
        run_timestamp = time.time()
        is_success, return_data = get_file_path(config_files_dir, main_config_id, "send_excel")
        attachment_dir = os.path.join(config_files_dir, str(main_config_id), "attachment")
        attachment_list = []
        if os.path.exists(attachment_dir):
            for attachment_name in os.listdir(attachment_dir):
                attachment_path = os.path.join(attachment_dir, attachment_name)
                attachment_list.append(attachment_path)
        history_list = []
        excel_data_list = []
        if not is_success:
            generate_log(main_config_id, "info", f"main_config_id {main_config_id} 发送无拆分附件邮件")
            for key, email_list in match_dict.items():
                single_history_list, single_excel_data_list = safe_send(smtp_obj, email_list, sender, subject, content, ip, port, password, key, main_config_id, run_timestamp, attachment_list)
                history_list.extend(single_history_list)
                excel_data_list.extend(single_excel_data_list)
        else:
            send_excel_path = return_data
            send_excel_name = os.path.splitext(os.path.split(send_excel_path)[1])[0]
            split_dir = os.path.join(config_files_dir, str(main_config_id), "split_excel")
            if os.path.exists(split_dir):
                shutil.rmtree(split_dir)
            os.makedirs(split_dir)
            if split_type == 1:
                send_workbook = openpyxl.load_workbook(send_excel_path, data_only=True)
                sheet_name_list = send_workbook.sheetnames
                target_list = sheet_name_list
                send_workbook.close()
                for sheet_name in sheet_name_list:
                    if sheet_name in match_dict:
                        split_excel_name = f'{clean_file_name(sheet_name)}_{send_excel_name}_.xlsx'
                        split_excel_path = os.path.join(split_dir, split_excel_name)
                        send_workbook = openpyxl.load_workbook(send_excel_path, data_only=True)
                        for delete_sheet_name in sheet_name_list:
                            if delete_sheet_name != sheet_name:
                                send_workbook.remove(send_workbook[delete_sheet_name])
                        send_workbook.save(split_excel_path)
                        send_workbook.close()
            else:
                empty_template_path = os.path.join(split_dir, "empty_template.xlsx")
                shutil.copy(send_excel_path, empty_template_path)
                generate_log(main_config_id, "info", f"main_config_id {main_config_id} 开始生成空模板")
                empty_excel = openpyxl.open(empty_template_path)
                empty_excel_sheet_name_list = empty_excel.sheetnames
                for empty_excel_sheet_name in empty_excel_sheet_name_list:
                    if empty_excel_sheet_name not in sheet_list:
                        empty_excel.remove(empty_excel[empty_excel_sheet_name])
                for sheet_name, field_row, is_split in zip(sheet_list, field_row_list, is_split_list):
                    empty_sheet = empty_excel[sheet_name]
                    header_row = get_header_row(empty_sheet, field_row)
                    max_row = int(empty_sheet.max_row)
                    generate_log(main_config_id, "info", f"main_config_id {main_config_id} sheet_name {sheet_name} max_row {max_row}")
                    if is_split:
                        empty_sheet.delete_rows(header_row+1, max_row+1)
                empty_excel.save(empty_template_path)
                empty_excel.close()
                generate_log(main_config_id, "info", f"main_config_id {main_config_id} 空模板表生成完毕")
                send_excel = openpyxl.load_workbook(send_excel_path, data_only=True)
                result_dict = {}
                target_list = []
                if split_field:
                    for sheet, field_row, is_split in zip(sheet_list, field_row_list, is_split_list):
                        send_sheet = send_excel[sheet]
                        field_row = get_header_row(send_sheet, field_row)
                        field_row = int(field_row) - 1
                        total_data_list = list(send_sheet.values)
                        data_list = total_data_list[field_row + 1:]
                        field_data_list = total_data_list[field_row]
                        split_field_index = field_data_list.index(split_field)
                        for data in data_list:
                            split_value = data[split_field_index]
                            if split_value not in target_list:
                                target_list.append(split_value)
                else:
                    target_list = list(match_dict.keys())
                for target in target_list:
                    result_dict[target] = {}
                    for sheet_name in sheet_list:
                        result_dict[target][sheet_name] = None
                generate_log(main_config_id, "info", f"main_config_id {main_config_id} 数据模板生成完毕")
                sheet_split_dict = dict(zip(sheet_list, is_split_list))
                for sheet_name, field_row, is_split in zip(sheet_list, field_row_list, is_split_list):
                    send_sheet = send_excel[sheet_name]
                    new_data_dict = {}
                    if is_split:
                        field_row = get_header_row(send_sheet, field_row)
                        field_row = int(field_row) - 1
                        total_data_list = list(send_sheet.values)
                        data_list = total_data_list[field_row+1:]
                        field_data_list = total_data_list[field_row]
                        split_field_index = field_data_list.index(split_field)
                        for data in data_list:
                            split_value = data[split_field_index]
                            if split_value not in new_data_dict:
                                new_data_dict[split_value] = []
                            new_data_dict[split_value].append(data)
                        for key, value in new_data_dict.items():
                            result_dict[key][sheet_name] = list(value)
                    else:
                        for target in target_list:
                            result_dict[target][sheet_name] = []
                send_excel.close()
                generate_log(main_config_id, "info", f"main_config_id {main_config_id} 数据生成完毕")
                for key, value in result_dict.items():
                    if key not in total_target_list:
                        generate_log(main_config_id, "error", f"main_config_id {main_config_id} {key} 未配置邮箱")
                        continue
                    single_excel_path = os.path.join(split_dir, f'{clean_file_name(key)}_{send_excel_name}_.xlsx')
                    shutil.copy(empty_template_path, single_excel_path)
                    work_book = openpyxl.load_workbook(single_excel_path, data_only=True)
                    for sheet_name, sheet_value in value.items():
                        is_split = sheet_split_dict[sheet_name]
                        if not is_split:
                            continue
                        work_sheet = work_book[sheet_name]
                        if sheet_value == None:
                            work_book.remove(work_sheet)
                            continue
                        for row_value in sheet_value:
                            work_sheet.append(row_value)
                    work_book.save(single_excel_path)
                    work_book.close()
                    generate_log(main_config_id, "info", f"main_config_id {main_config_id} {key} 表生成完毕")
                os.remove(empty_template_path)
            # smtp_obj.quit()
            smtp_obj = smtplib.SMTP(ip, port)
            smtp_obj.login(sender, password)
            for key in target_list:
                if key in match_dict:
                    email_list = match_dict[key]
                    split_path = os.path.join(split_dir, f'{clean_file_name(key)}_{send_excel_name}_.xlsx')
                    all_attachment_list = [split_path] + attachment_list
                    single_history_list, single_excel_data_list = safe_send(smtp_obj, email_list, sender, subject, content, ip, port, password, key, main_config_id, run_timestamp, all_attachment_list)
                    history_list.extend(single_history_list)
                    excel_data_list.extend(single_excel_data_list)
        session = Session()
        session.query(SendHistory).filter_by(main_config_id=main_config_id).delete()
        session.add_all(history_list)
        session.query(MainConfig).filter_by(id=main_config_id).update({"run_timestamp": run_timestamp})
        session.commit()
        session.close()
        workbook = openpyxl.Workbook()
        sheet = workbook[workbook.sheetnames[0]]
        sheet.append(["序号", "单位", "邮箱", "状态"])
        for data_index, excel_data in enumerate(excel_data_list, start=1):
            sheet.append([data_index]+excel_data)
        status_excel_path = os.path.join(config_files_dir, str(main_config_id), "发件状态表.xlsx")
        workbook.save(status_excel_path)
        workbook.close()
        generate_log(main_config_id, "info", f"main_config_id {main_config_id}  运行成功")
        # smtp_obj.quit()
        return "运行成功"
    except Exception as error:
        generate_log(main_config_id, "critical", f"main_config_id {main_config_id} 运行失败 {print_exc()}")
        return "运行失败,未知错误"


