import os
import openpyxl
from app.parameter_config import check_file_dict
from .public_func import generate_key_value, replace_company, generate_dict, generate_change_info, generate_percent_rate

def compare_with_last(data_dict, last_file_path):
    if os.path.exists(last_file_path):
        workbook = openpyxl.load_workbook(last_file_path)
        sheet = workbook[workbook.sheetnames[0]]
        last_data_dict = generate_dict(sheet, 1, ["公司"], ["本月成本挂账总额", "本月挂账2年以上金额"])
        workbook.close()
    else:
        last_data_dict = {}
    result_list = []
    for company, data in data_dict.items():
        last_data = last_data_dict.get(company, (0, 0))
        amount, year_amount = data
        last_amount, last_year_amount = [float(i) for i in last_data]
        handle_amount, change_amount, rate = generate_change_info(amount, last_amount)
        handle_year_amount, change_year_amount, year_rate = generate_change_info(year_amount, last_year_amount)
        result_list.append([company, handle_amount, last_amount, change_amount, rate, handle_year_amount, last_year_amount, change_year_amount, year_rate])
    return result_list

def generate_result_list(file_path, centre_company_dict, last_file_path, code_department_dict):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[workbook.sheetnames[0]]
    column_value_list = list(sheet.columns)
    field_list = [i.value for i in sheet[1]]
    data_list = generate_key_value(column_value_list, field_list, ["公司代码", "利润中心", "金额", "2-3年", "3-4年", "4-5年", "5年以上"], 1)
    workbook.close()
    for data_index, data in enumerate(data_list[:]):
        code = data[0]
        department = code_department_dict.get(code)
        data_list[data_index].insert(1, department)
    replace_company(data_list, centre_company_dict)
    data_dict = {}
    for data in data_list:
        company = data[1]
        amount = float(data[3]) if data[3] else 0
        year_amount = sum(float(i) if i else 0 for i in data[4:])
        data_dict[company] = [data_dict[company][0] + amount, data_dict[company][1] + year_amount] if company in data_dict else [amount, year_amount]
    return compare_with_last(data_dict, last_file_path)

def generate_file(result_list, result_file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook[workbook.sheetnames[0]]
    sheet.append(["序号", "公司", "本月成本挂账总额", "年初成本挂账总额", "总额较年初增减额", "总额较年初变化幅度(%)", "本月挂账2年以上金额", "年初挂账2年以上金额", "挂账两年以上较年初增减额", "挂账2年以上较年初变化幅度(%)"])
    result_list.sort(key=lambda x:x[5])
    sum_amount = 0
    sum_last_amount = 0
    sum_year_amount = 0
    sum_last_year_amount = 0
    for result_index, result in enumerate(result_list, start=1):
        sum_amount += result[1]
        sum_last_amount += result[2]
        sum_year_amount += result[5]
        sum_last_year_amount += result[6]
        sheet.append([result_index]+result)
    sum_change_amount = sum_amount - sum_last_amount
    sum_change_year_amount = sum_year_amount - sum_last_year_amount
    sum_rate = generate_percent_rate(sum_change_amount, sum_last_amount)
    sum_year_rate = generate_percent_rate(sum_change_year_amount, sum_last_year_amount)
    sheet.append([None, "合计", sum_amount, sum_last_amount, sum_change_amount, sum_rate, sum_year_amount, sum_last_year_amount, sum_change_year_amount, sum_year_rate])
    workbook.save(result_file_path)
    workbook.close()

def handle_6(file_dir, last_file_dir, centre_company_dict, code_department_dict):
    file_name = "生产成本长期挂账未结转情况统计表.xlsx"
    origin_file_dir = os.path.join(file_dir, "origin")
    file_path = os.path.join(origin_file_dir, f"{check_file_dict['product_cost_excel']}.xlsx")
    if not os.path.exists(file_path):
        print(f"6文件丢失")
        return
    last_file_path = os.path.join(last_file_dir, file_name)
    result_list = generate_result_list(file_path, centre_company_dict, last_file_path, code_department_dict)
    result_file_dir = os.path.join(file_dir, "result")
    if not os.path.exists(result_file_dir):
        os.makedirs(result_file_dir)
    result_file_path = os.path.join(result_file_dir, file_name)
    generate_file(result_list, result_file_path)
    print("6run end")