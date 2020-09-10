import os
import openpyxl

class Report:
    def __init__(self, result_workbook, filter_workbook, file_dir, file_name, header_row, sheet_name, field_list, result_field_list, merge_field_list, centre_company_dict, code_department_dict, last_file_dir=None, company_index=1, is_save_all=True, prefix=''):
        self.sheet_name = sheet_name
        self.file_dir = file_dir
        self.origin_file_path = os.path.join(file_dir, "origin", file_name)
        self.result_workbook = result_workbook
        self.filter_workbook = filter_workbook
        self.filter_sheet_name = prefix + file_name.split(".")[0]
        self.last_file_path = os.path.join(last_file_dir, "稽核处结果表.xlsx") if last_file_dir else None
        self.field_list = field_list
        self.header_row = header_row
        self.centre_company_dict = centre_company_dict
        self.code_department_dict = code_department_dict
        self.merge_field_list = merge_field_list
        self.result_field_list = result_field_list
        self.save_company_list = []
        self.company_index = company_index
        self.is_save_all = is_save_all

    @staticmethod
    def generate_key_value(column_value_list, field_list, target_field_list, header_row):
        target_list = []
        for target_field in target_field_list:
            target_index = field_list.index(target_field)
            target_list.append([i.value for i in column_value_list[target_index][header_row:]])
        if len(target_list) == 1:
            target = [str(i) if i else i for i in target_list[0]]
        else:
            target = [[str(j) if j else j for j in i] for i in zip(*target_list)]
        return target

    @staticmethod
    def generate_dict(sheet, header_row, key_field_list, value_field_list):
        field_list = [i.value for i in sheet[header_row]]
        column_value_list = list(sheet.columns)
        key_list = Report.generate_key_value(column_value_list, field_list, key_field_list, header_row)
        value_list = Report.generate_key_value(column_value_list, field_list, value_field_list, header_row)
        data_dict = dict(zip(key_list, value_list))
        data_dict.pop(None, None)
        return data_dict

    @staticmethod
    def handle_num(num):
        return round(float(num) / 10000, 2)

    @staticmethod
    def generate_change_info(amount, last_amount):
        handle_amount = Report.handle_num(amount)
        change_amount = handle_amount - float(last_amount)
        rate = Report.generate_percent_rate(change_amount, last_amount)
        return handle_amount, change_amount, rate

    @staticmethod
    def generate_percent_rate(num1, num2):
        return f'{round((float(num1) / float(num2)) * 100, 2)}%' if float(num2) else None

    def replace_company(self, data_list):
        remove_data_list = []
        for data_index, data in enumerate(data_list[:]):
            if not data[1]:
                remove_data_list.append(data)
                continue
            if data[0] in ["4600", "4606", "4608", "4609"]:
                if data[2] not in self.centre_company_dict:
                    print(f"缺失 {data[2]}")
                    remove_data_list.append(data)
                    continue
                data_list[data_index][1] = self.centre_company_dict[data[2]]
        for remove_data in remove_data_list:
            data_list.remove(remove_data)


    def get_data_list(self):
        workbook = openpyxl.load_workbook(self.origin_file_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        column_value_list = list(sheet.columns)
        all_field_list = [i.value for i in sheet[self.header_row]]
        order_field_list = list(self.field_list)
        for field in all_field_list:
            if field not in self.field_list:
                order_field_list.append(field)
        data_list = Report.generate_key_value(column_value_list, all_field_list, order_field_list, self.header_row)
        workbook.close()
        return data_list, order_field_list

    def fix_data_list(self, data_list, order_field_list):
        pass

    def merge_data(self, data_list):
        data_dict = {}
        return data_dict

    def get_last_data(self):
        if os.path.exists(self.last_file_path):
            print("旧数据存在")
            last_excel = openpyxl.load_workbook(self.last_file_path, data_only=True)
            last_sheet = last_excel[self.sheet_name]
            last_data_dict = Report.generate_dict(last_sheet, 1, ["公司名称"], self.merge_field_list)
            last_excel.close()
        else:
            print("旧数据不存在")
            last_data_dict = {}
        return last_data_dict

    def get_result_list(self, data_dict, count_dict):
        return []

    def save(self, result_list):
        sheet = self.result_workbook.create_sheet(self.sheet_name, -1)
        for result in result_list:
            sheet.append(result)

    def filter_save(self, data_list, order_field_list):
        sheet = self.filter_workbook.create_sheet(self.filter_sheet_name, -1)
        sheet.append(order_field_list)
        if self.save_company_list:
            for data in data_list:
                if data[self.company_index] in self.save_company_list:
                    sheet.append(data)

    def start(self):
        print("-"*20)
        print(self.sheet_name)
        print("开始")
        if not os.path.exists(self.origin_file_path):
            print(f"原文件不存在")
            return
        data_list, order_field_list = self.get_data_list()
        filter_data_list = self.fix_data_list(data_list, order_field_list)
        self.replace_company(data_list)
        data_dict, count_dict = self.merge_data(data_list)
        result_list = self.get_result_list(data_dict, count_dict)
        result_list.insert(0, self.result_field_list)
        print("生成统计表")
        self.save(result_list)
        print("生成过滤表")
        if filter_data_list:
            self.replace_company(filter_data_list)
            self.filter_save(filter_data_list, order_field_list)
        else:
            self.filter_save(data_list, order_field_list)
        print("结束")
        print("-" * 20)