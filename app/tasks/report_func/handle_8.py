import os
import openpyxl
from app.parameter_config import check_file_dict
from .public_func import generate_key_value, replace_company, generate_dict, generate_change_info, generate_percent_rate, get_data_list

def compare_with_last(data_dict, last_file_path):
    if os.path.exists(last_file_path):
        workbook = openpyxl.load_workbook(last_file_path)
        sheet = workbook[workbook.sheetnames[0]]
        last_data_dict = generate_dict(sheet, 1, ["单位名称"], ["本月合计金额", "本月1年以上金额"])
        workbook.close()
    else:
        last_data_dict = {}
    result_list = []
    for company, data in data_dict.items():
        last_data = last_data_dict.get(company, (0, 0))
        amount, year_amount = data
        last_amount, last_year_amount = [float(i) for i in last_data]
        handle_amount, change_amount, rate = generate_change_info(amount, last_amount)
        handle_year_amount, change_year_amount, year_rate = generate_change_info(year_amount, last_year_amount)
        result_list.append([company, handle_amount, last_amount, change_amount, rate, handle_year_amount, last_year_amount, change_year_amount, year_rate])
    return result_list

def generate_result_list(file_path, centre_company_dict, last_file_path, code_department_dict):
    data_list = get_data_list(file_path, 1, 1, ["利润中心", "金额", "1-2年", "2-3年", "3年以上"])
    for data_index, data in enumerate(data_list[:]):
        centre = data[0]
        code = centre[1:5]
        department = code_department_dict.get(code)
        data_list[data_index].insert(0, department)
        data_list[data_index].insert(0, code)
    replace_company(data_list, centre_company_dict)
    data_dict = {}
    for data in data_list:
        company = data[1]
        amount = float(data[3])
        year_amount = sum(float(i) if i else 0 for i in data[4:])
        data_dict[company] = [data_dict[company][0] + amount, data_dict[company][1] + year_amount] if company in data_dict else [amount, year_amount]
    return compare_with_last(data_dict, last_file_path)

def generate_file(result_list, result_file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook[workbook.sheetnames[0]]
    sheet.append(["序号", "单位名称", "本月合计金额", "年初合计金额", "合计金额增减变化", "合计金额增降幅 %", "本月1年以上金额", "年初1年以上金额", "1年以上金额增长额", "1年以上金额增降幅 %"])
    result_list.sort(key=lambda x:x[5])
    sum_amount = 0
    sum_last_amount = 0
    sum_year_amount = 0
    sum_last_year_amount = 0
    for result_index, result in enumerate(result_list, start=1):
        sum_amount += result[1]
        sum_last_amount += result[2]
        sum_year_amount += result[5]
        sum_last_year_amount += result[6]
        sheet.append([result_index]+result)
    sum_change_amount = sum_amount - sum_last_amount
    sum_change_year_amount = sum_year_amount - sum_last_year_amount
    sum_rate = generate_percent_rate(sum_change_amount, sum_last_amount)
    sum_year_rate = generate_percent_rate(sum_change_year_amount, sum_last_year_amount)
    sheet.append([None, "合计", sum_amount, sum_last_amount, sum_change_amount, sum_rate, sum_year_amount, sum_last_year_amount, sum_change_year_amount, sum_year_rate])
    workbook.save(result_file_path)
    workbook.close()

def handle_8(file_dir, last_file_dir, centre_company_dict, code_department_dict):
    file_name = "挂账一年以上应付原材料暂估情况分析.xlsx"
    origin_file_dir = os.path.join(file_dir, "origin")
    file_path = os.path.join(origin_file_dir, f"{check_file_dict['pay_receive_excel']}.xlsx")
    if not os.path.exists(file_path):
        print(f"8文件丢失")
        return
    last_file_path = os.path.join(last_file_dir, file_name)
    result_list = generate_result_list(file_path, centre_company_dict, last_file_path, code_department_dict)
    result_file_dir = os.path.join(file_dir, "result")
    if not os.path.exists(result_file_dir):
        os.makedirs(result_file_dir)
    result_file_path = os.path.join(result_file_dir, file_name)
    generate_file(result_list, result_file_path)
    print("8run end")