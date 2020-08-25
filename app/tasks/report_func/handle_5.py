import os
import openpyxl
from datetime import datetime, timedelta
from app.parameter_config import check_file_dict
from .public_func import generate_key_value, replace_company, handle_num

def generate_result_list(file_path, centre_company_dict, last_date):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[workbook.sheetnames[0]]
    column_value_list = list(sheet.columns)
    field_list = [i.value for i in sheet[4]]
    data_list = generate_key_value(column_value_list, field_list, ["公司代码", "公司名称", "利润中心", "在制品余额", "项目最近确认收入日期", "项目个数"], 4)
    workbook.close()
    replace_company(data_list, centre_company_dict)
    data_dict = {}
    count_dict = {}
    for data in data_list:
        data_date_str = data[4]
        if not data_date_str:
            continue
        data_date = datetime.strptime(str(data_date_str), "%Y%m%d")
        if data_date >= last_date:
            print(data_date)
            continue
        company = data[1]
        amount = float(data[3])
        count = int(data[5])
        data_dict[company] = data_dict[company] + amount if company in data_dict else amount
        count_dict[company] = count_dict[company] + count if company in count_dict else count
    result_list = []
    for company, amount in data_dict.items():
        count = count_dict[company]
        handle_amount = handle_num(amount)
        result_list.append([company, handle_amount, count])
    return result_list

def generate_file(result_list, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook[workbook.sheetnames[0]]
    result_list.sort(key=lambda x:x[1], reverse=True)
    sheet.append(["序号", "公司名称", "在制品余额 万元", "项目个数 个"])
    sum_amount = 0
    sum_count = 0
    for result_index, result in enumerate(result_list, start=1):
        sum_amount += result[1]
        sum_count += result[2]
        sheet.append([result_index]+result)
    sheet.append([None, "合计", sum_amount, sum_count])
    workbook.save(file_path)
    workbook.close()

def handle_5(file_dir, centre_company_dict):
    print("run start")
    file_name = "项目成本结转不彻底.xlsx"
    now_date = datetime.now()
    last_date = datetime(year=now_date.year, month=now_date.month, day=1)
    print(last_date)
    origin_file_dir = os.path.join(file_dir, "origin")
    file_path = os.path.join(origin_file_dir, f"{check_file_dict['balance_analyse_excel']}.xlsx")
    if not os.path.exists(file_path):
        print("文件丢失")
        return
    result_list = generate_result_list(file_path, centre_company_dict, last_date)
    result_file_dir = os.path.join(file_dir, "result")
    if not os.path.exists(result_file_dir):
        os.makedirs(result_file_dir)
    result_file_path = os.path.join(result_file_dir, file_name)
    generate_file(result_list, result_file_path)
    print("run end")