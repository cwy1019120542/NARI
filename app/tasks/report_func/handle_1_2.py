import os
import openpyxl
from .public_func import generate_dict, generate_key_value, handle_num, replace_company, generate_percent_rate, generate_change_info
from app.parameter_config import check_file_dict

def compare_with_last(data_dict, file_name, last_file_dir, result_amount_field):
    result_list = []
    if os.path.exists(last_file_dir) and file_name in os.listdir(last_file_dir):
        print("last_file_dir exists")
        last_excel = openpyxl.load_workbook(os.path.join(last_file_dir, file_name), data_only=True)
        last_sheet = last_excel[last_excel.sheetnames[0]]
        last_data_dict = generate_dict(last_sheet, 1, ["公司名称"], result_amount_field)
        last_excel.close()
    else:
        print("last_file_dir not exists")
        last_data_dict = {}
    for company, amount in data_dict.items():
        last_amount = float(last_data_dict.get(company, 0))
        handle_amount, change_amount, rate = generate_change_info(amount, last_amount)
        result_list.append([company, handle_amount, last_amount, change_amount, rate])
    return result_list

def generate_result_list(file_path, key_word, centre_company_dict, total_file_name, inner_file_name, last_file_dir, result_amount_field):
    excel = openpyxl.load_workbook(file_path, data_only=True)
    sheet = excel[excel.sheetnames[0]]
    column_value_list = list(sheet.columns)
    field_list = [i.value for i in sheet[1]]
    data_list = generate_key_value(column_value_list, field_list, ["公司代码", "公司名称", "利润中心", f"{key_word}开票金额", "客户属性"], 1)
    excel.close()
    replace_company(data_list, centre_company_dict)
    total_data_dict = {}
    inner_data_dict = {}
    for data in data_list:
        company = data[1]
        amount = float(data[3]) if data[3] else 0
        if data[4] == "国网系统内-集团内":
            inner_data_dict[company] = inner_data_dict[company] + amount if company in inner_data_dict else amount
        total_data_dict[company] = total_data_dict[company] + amount if company in total_data_dict else amount
    print("start compare_with_last total")
    total_result_list = compare_with_last(total_data_dict, total_file_name, last_file_dir, result_amount_field)
    print("end compare_with_last total")
    print("start compare_with_last inner")
    inner_result_list = compare_with_last(inner_data_dict, inner_file_name, last_file_dir, result_amount_field)
    print("end compare_with_last inner")
    return total_result_list, inner_result_list

def generate_file(result_list, result_dir, file_name, key_word):
    file_path = os.path.join(result_dir, file_name)
    result_list.sort(key=lambda x:x[3], reverse=True)
    workbook = openpyxl.Workbook()
    sheet = workbook[workbook.sheetnames[0]]
    field_list = ["序号", "公司名称", f"本月{key_word}开票金额", f"上月{key_word}开票金额", f"新增{key_word}开票金额", "较上月增幅"]
    sheet.append(field_list)
    sum_amount = 0
    sum_last_amount = 0
    sum_change_amount = 0
    for result_index, result in enumerate(result_list, start=1):
        sum_amount += result[1]
        sum_last_amount += result[2]
        sum_change_amount += result[3]
        sheet.append([result_index] + result)
    sum_rate = generate_percent_rate(sum_change_amount, sum_last_amount)
    sum_list = [None, "合计", sum_amount, sum_last_amount, sum_change_amount, sum_rate]
    sheet.append(sum_list)
    workbook.save(file_path)
    workbook.close()

def handle_1_2(file_dir, last_file_dir, centre_company_dict):
    total_file1_name = "本月新增预开票情况统计表(全部).xlsx"
    total_file2_name = "本月新增滞后开票情况统计表(全部).xlsx"
    inner_file1_name = "本月新增预开票情况统计表(系统内).xlsx"
    inner_file2_name = "本月新增滞后开票情况统计表(系统内).xlsx"
    file1_key_word = "预"
    file2_key_word = "滞后"
    origin_file_dir = os.path.join(file_dir, "origin")
    file1_path = os.path.join(origin_file_dir, f"{check_file_dict['pre_excel']}.xlsx")
    file2_path = os.path.join(origin_file_dir, f"{check_file_dict['suf_excel']}.xlsx")
    if not os.path.exists(file1_path) or not os.path.exists(file2_path):
        print("文件丢失")
        return
    print("start generate_result_dict 预开票金额")
    total_result_list1, inner_result_list1 = generate_result_list(file1_path, file1_key_word, centre_company_dict, total_file1_name, inner_file1_name, last_file_dir, ["本月预开票金额"])
    print("end generate_result_dict 预开票金额")
    print("start generate_result_dict 滞后开票金额")
    total_result_list2, inner_result_list2 = generate_result_list(file2_path, file2_key_word, centre_company_dict, total_file2_name, inner_file2_name, last_file_dir, ["本月滞后开票金额"])
    print("end generate_result_dict 滞后开票金额")
    print("start generate_file total")
    result_file_dir = os.path.join(file_dir, "result")
    if not os.path.exists(result_file_dir):
        os.makedirs(result_file_dir)
    generate_file(total_result_list1, result_file_dir, total_file1_name, file1_key_word)
    generate_file(inner_result_list1, result_file_dir, inner_file1_name, file1_key_word)
    generate_file(total_result_list2, result_file_dir, total_file2_name, file2_key_word)
    generate_file(inner_result_list2, result_file_dir, inner_file2_name, file2_key_word)
    print("运行结束")