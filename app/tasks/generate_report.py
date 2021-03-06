import os
import shutil
import openpyxl
from ..extention import celery
from .report_func.all_report import *
from ..parameter_config import check_file_dict
from ..func_tools import generate_report_word

@celery.task
def generate_report(file_dir, code_file_path, last_file_dir, last_year_file_dir, date):
    workbook = openpyxl.load_workbook(code_file_path, data_only=True)
    code_sheet = workbook[workbook.sheetnames[0]]
    code_department_dict = Report.generate_dict(code_sheet, 1, ["公司代码"], ["单位名称"])
    centre_sheet = workbook[workbook.sheetnames[1]]
    centre_company_dict = Report.generate_dict(centre_sheet, 1, ["利润中心"], ["分公司及事业部名称"])
    manage_sheet = workbook[workbook.sheetnames[2]]
    manage_dict = Report.generate_dict(manage_sheet, 1, ["匹配元素"], ["是否归属本部管理（是/否）"])
    workbook.close()
    result_dir = os.path.join(file_dir, "result")
    if os.path.exists(result_dir):
        shutil.rmtree(result_dir)
    os.makedirs(result_dir)
    result_workbook = openpyxl.Workbook()
    result_workbook.save(os.path.join(result_dir, "稽核处结果表.xlsx"))
    result_workbook.close()
    filter_workbook1 = openpyxl.Workbook()
    Report_2(filter_workbook1, file_dir, f"{check_file_dict['pre_excel']}.xlsx", 1, "本月新增预开票情况统计表(系统内)",
             ["公司代码", "公司名称", "利润中心", "预开票金额", "客户属性"], ["序号", "公司名称", "本月预开票金额", "上月预开票金额", "新增预开票金额", "较上月增幅"],
             ["本月预开票金额"], centre_company_dict, code_department_dict, last_file_dir=last_file_dir, prefix="(本月新增)(系统内)", manage_dict=manage_dict).start()
    Report_2(filter_workbook1, file_dir, f"{check_file_dict['suf_excel']}.xlsx", 1, "本月新增滞后开票情况统计表(系统内)",
             ["公司代码", "公司名称", "利润中心", "滞后开票金额", "客户属性"], ["序号", "公司名称", "本月滞后开票金额", "上月滞后开票金额", "新增滞后开票金额", "较上月增幅"],
             ["本月滞后开票金额"], centre_company_dict, code_department_dict, last_file_dir=last_file_dir, prefix="(本月新增)(系统内)", manage_dict=manage_dict).start()
    Report_1(filter_workbook1, file_dir, f"{check_file_dict['pre_excel']}.xlsx", 1, "本月新增预开票情况统计表(全部)",
             ["公司代码", "公司名称", "利润中心", "预开票金额", "客户属性"], ["序号", "公司名称", "本月预开票金额", "上月预开票金额", "新增预开票金额", "较上月增幅"],
             ["本月预开票金额"], centre_company_dict, code_department_dict, last_file_dir=last_file_dir,
             prefix="(本月新增)(全部)", manage_dict=manage_dict).start()
    Report_1(filter_workbook1, file_dir, f"{check_file_dict['suf_excel']}.xlsx", 1, "本月新增滞后开票情况统计表(全部)",
             ["公司代码", "公司名称", "利润中心", "滞后开票金额", "客户属性"], ["序号", "公司名称", "本月滞后开票金额", "上月滞后开票金额", "新增滞后开票金额", "较上月增幅"],
             ["本月滞后开票金额"], centre_company_dict, code_department_dict, last_file_dir=last_file_dir,
             prefix="(本月新增)(全部)", manage_dict=manage_dict).start()
    Report_4(filter_workbook1, file_dir, f"{check_file_dict['pre_excel']}.xlsx", 1, "已开票未确认收入(预开票)余额清理情况统计表(系统内)",
             ["公司代码", "公司名称", "利润中心", f"预开票金额", "客户属性", "1-2年", "2-3年", "3年以上"],
             ["序号", "公司名称", f"本月预开票金额", f"年初预开票金额", "总额较年初增减额", "总额较年初变化幅度(%)", "挂账1年以上金额", "年初挂账1年以上金额",
              "挂账1年以上较年初增减额", "挂账1年以上较年初变化幅度(%)"],
             ["本月预开票金额", "挂账1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, prefix="(余额清理情况)(系统内)", manage_dict=manage_dict).start()
    Report_4(filter_workbook1, file_dir, f"{check_file_dict['suf_excel']}.xlsx", 1, "已开票未确认收入(滞后开票)余额清理情况统计表(系统内)",
             ["公司代码", "公司名称", "利润中心", f"滞后开票金额", "客户属性", "1-2年", "2-3年", "3年以上"],
             ["序号", "公司名称", "本月滞后开票金额", f"年初滞后开票金额", "总额较年初增减额", "总额较年初变化幅度(%)", "挂账1年以上金额", "年初挂账1年以上金额",
              "挂账1年以上较年初增减额", "挂账1年以上较年初变化幅度(%)"],
             ["本月滞后开票金额", "挂账1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, prefix="(余额清理情况)(系统内)", manage_dict=manage_dict).start()
    Report_3(filter_workbook1, file_dir, f"{check_file_dict['pre_excel']}.xlsx", 1, "已开票未确认收入(预开票)余额清理情况统计表(全部)",
             ["公司代码", "公司名称", "利润中心", f"预开票金额", "客户属性", "1-2年", "2-3年", "3年以上"],
             ["序号", "公司名称", f"本月预开票金额", f"年初预开票金额", "总额较年初增减额", "总额较年初变化幅度(%)", "挂账1年以上金额", "年初挂账1年以上金额",
              "挂账1年以上较年初增减额", "挂账1年以上较年初变化幅度(%)"],
             ["本月预开票金额", "挂账1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, prefix="(余额清理情况)(全部)", manage_dict=manage_dict).start()
    Report_3(filter_workbook1, file_dir, f"{check_file_dict['suf_excel']}.xlsx", 1, "已开票未确认收入(滞后开票)余额清理情况统计表(全部)",
             ["公司代码", "公司名称", "利润中心", f"滞后开票金额", "客户属性", "1-2年", "2-3年", "3年以上"],
             ["序号", "公司名称", "本月滞后开票金额", f"年初滞后开票金额", "总额较年初增减额", "总额较年初变化幅度(%)", "挂账1年以上金额", "年初挂账1年以上金额",
              "挂账1年以上较年初增减额", "挂账1年以上较年初变化幅度(%)"],
             ["本月滞后开票金额", "挂账1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, prefix="(余额清理情况)(全部)", manage_dict=manage_dict).start()
    Report_5(filter_workbook1, file_dir, f"{check_file_dict['balance_analyse_excel']}.xlsx", 4, "项目成本结转不彻底",
             ["公司代码", "公司名称", "利润中心", "在制品余额", "项目最近确认收入日期", "项目个数"],
             ["序号", "公司名称", "在制品余额 万元", "项目个数 个"],
             [], centre_company_dict, code_department_dict,
             last_file_dir=None, manage_dict=manage_dict).start()
    Report_6(filter_workbook1, file_dir, f"{check_file_dict['product_cost_excel']}.xlsx", 1, "生产成本长期挂账未结转情况统计表",
             ["公司代码", "利润中心", "金额", "2-3年", "3-4年", "4-5年", "5年以上"],
             ["序号", "公司名称", "本月成本挂账总额", "年初成本挂账总额", "总额较年初增减额", "总额较年初变化幅度(%)", "本月挂账2年以上金额", "年初挂账2年以上金额", "挂账两年以上较年初增减额", "挂账2年以上较年初变化幅度(%)"],
             ["本月成本挂账总额", "本月挂账2年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, is_save_all=False, manage_dict=manage_dict).start()
    Report_7(filter_workbook1, file_dir, f"{check_file_dict['pay_cost_excel']}.xlsx", 1, "挂账一年以上应付项目暂估情况分析",
             ["利润中心", "金额", "1-2年", "2-3年", "3年以上"],
             ["序号", "公司名称", "本月合计金额", "年初合计金额", "合计金额增减变化", "合计金额增降幅 %", "本月1年以上金额", "年初1年以上金额", "1年以上金额增长额", "1年以上金额增降幅 %"],
             ["本月合计金额", "本月1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, manage_dict=manage_dict).start()
    filter_workbook1.save(os.path.join(result_dir, "稽核处中间表1.xlsx"))
    filter_workbook1.close()
    filter_workbook2 = openpyxl.Workbook()
    Report_8(filter_workbook2, file_dir, f"{check_file_dict['pay_receive_excel']}.xlsx", 1, "挂账一年以上应付原材料暂估情况分析",
             ["利润中心", "金额", "1-2年", "2-3年", "3年以上"],
             ["序号", "公司名称", "本月合计金额", "年初合计金额", "合计金额增减变化", "合计金额增降幅 %", "本月1年以上金额", "年初1年以上金额", "1年以上金额增长额", "1年以上金额增降幅 %"],
             ["本月合计金额", "本月1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, manage_dict=manage_dict).start()
    filter_workbook2.save(os.path.join(result_dir, "稽核处中间表2.xlsx"))
    filter_workbook2.close()
    filter_workbook3 = openpyxl.Workbook()
    Report_9(filter_workbook3, file_dir, f"{check_file_dict['pre_pay_excel']}.xlsx", 1, "挂账一年以上预付账款情况",
             ["利润中心", "金额", "1-2年", "2-3年", "3-4年", "4-5年", "5年以上"],
             ["序号", "公司名称", "本月合计金额", "年初合计金额", "合计金额增减变化", "合计金额增降幅 %", "本月1年以上金额", "年初1年以上金额", "1年以上金额增长额", "1年以上金额增降幅 %"],
             ["本月合计金额", "本月1年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, manage_dict=manage_dict).start()
    Report_10(filter_workbook3, file_dir, f"{check_file_dict['other_receive_excel']}.xlsx", 1, "挂账三年以上其他应收账款情况",
             ["利润中心", "本币金额", "3-4年", "4-5年", "5年以上"],
             ["序号", "公司名称", "本月合计金额", "年初合计金额", "合计金额增减变化", "合计金额增降幅 %", "本月3年以上金额", "年初3年以上金额", "3年以上金额增长额", "3年以上金额增降幅 %"],
             ["本月合计金额", "本月3年以上金额"], centre_company_dict, code_department_dict,
             last_file_dir=last_year_file_dir, manage_dict=manage_dict).start()
    Report_11(filter_workbook3, file_dir, f"{check_file_dict['other_pay_excel']}.xlsx", 1, "挂账三年以上其他应付账款情况",
              ["利润中心", "余额", "3-4年", "4-5年", "5年以上"],
              ["序号", "公司名称", "本月合计金额", "年初合计金额", "合计金额增减变化", "合计金额增降幅 %", "本月3年以上金额", "年初3年以上金额", "3年以上金额增长额", "3年以上金额增降幅 %"],
              ["本月合计金额", "本月3年以上金额"], centre_company_dict, code_department_dict,
              last_file_dir=last_year_file_dir, manage_dict=manage_dict).start()
    Report_12(filter_workbook3, file_dir, f"{check_file_dict['receive_excel']}.xlsx", 2, "内部关联交易-收入确认与收货不同步",
              ["单位名称", "销售确认收入金额", "采购收货金额"],
              ["序号", "公司名称", "销售确认收入金额", "采购收货金额", "差异金额", "个数"],
              [], centre_company_dict, code_department_dict, None, 0, manage_dict=manage_dict).start()
    Report_13(filter_workbook3, file_dir, f"{check_file_dict['open_excel']}.xlsx", 1, "内部关联交易-收入确认与开票不同步",
              ["公司代码", "公司名称", "利润中心", "客户名称", "销售方开票金额", "确认收入金额", "是否一致"],
              ["序号", "公司名称", "预开票(金额)", "预开票(个数)", "滞后开票(金额)", "滞后开票(个数)"],
              [], centre_company_dict, code_department_dict,
              last_file_dir=None, manage_dict=manage_dict).start()
    Report_14(filter_workbook3, file_dir, f"{check_file_dict['error_excel']}.xlsx", 4, "本月项目生产成本暂估比例异常(达20%以上)情况统计表",
              ["公司代码", "公司名称", "利润中心", "累计：生产成本暂估", "暂估占结转成本比(%)"],
              ["序号", "公司名称", "本月异常生产成本暂估金额 万元", "本月异常的项目数量 个",
               "年初异常生产成本暂估金额 万元", "年初异常的项目数量 个", "异常生产成本暂估较年初金额增减变化 万元", "较年初增降幅度 %"],
              ["本月异常生产成本暂估金额 万元", "本月异常的项目数量 个"], centre_company_dict, code_department_dict,
              last_file_dir=last_year_file_dir, manage_dict=manage_dict).start()
    filter_workbook3.save(os.path.join(result_dir, "稽核处中间表3.xlsx"))
    filter_workbook3.close()
    file_path = os.path.join(result_dir, "稽核处结果表.xlsx")
    result_path = os.path.join(result_dir, "报告.docx")
    generate_report_word(5, file_path, date, result_path)
    print("保存完成")

